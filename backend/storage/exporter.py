"""Exporter and Importer module for lmScraper events."""
import pandas as pd
import io
import aiofiles
from pathlib import Path
from datetime import datetime
from fastapi.responses import StreamingResponse
from backend.storage.db import (
    query_events, get_events_by_ids, get_all_event_ids,
    bulk_update_events, insert_event, IMPORT_EDITABLE_FIELDS,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

EXPORT_COLS = [
    'record_id', 'hidden',
    'title', 'date_start', 'date_end', 'city', 'country', 'venue', 'price',
    'event_type', 'dance_style', 'organizer_name', 'organizer_email',
    'organizer_phone', 'organizer_instagram', 'organizer_facebook',
    'organizer_tiktok', 'organizer_whatsapp', 'organizer_youtube',
    'organizer_twitter', 'organizer_website', 'contact_hidden',
    'event_url', 'source_domain', 'platform', 'scraped_at',
]


def _normalize_event_date(val):
    if not val or (isinstance(val, float) and pd.isna(val)):
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    try:
        if len(val_str) >= 10 and val_str[4] == '-' and val_str[7] == '-':
            dt = pd.to_datetime(val_str)
            return dt.strftime("%d/%m/%Y %H:%M")
        elif '/' in val_str:
            parts = val_str.split()
            date_part = parts[0]
            time_part = parts[1] if len(parts) > 1 else '00:00'
            subparts = date_part.split('/')
            if len(subparts) == 3:
                d, m, y = subparts
                return f"{int(d):02d}/{int(m):02d}/{int(y):04d} {time_part[:5]}"
        return pd.to_datetime(val_str, dayfirst=True).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return val_str

_format_iso_date = _normalize_event_date


async def get_events_df(filters: dict) -> pd.DataFrame:
    """Get all events matching *filters* as a DataFrame ready for export."""
    events, _ = await query_events(filters, page=1, per_page=999999)
    if not events:
        return pd.DataFrame()

    events_export = []
    for event in events:
        e = dict(event)
        e['record_id'] = e.get('id')
        e['event_type'] = e.get('category')
        e['hidden'] = 'yes' if e.get('is_hidden') else 'no'
        events_export.append(e)

    df = pd.DataFrame(events_export)
    for col in EXPORT_COLS:
        if col not in df.columns:
            df[col] = ''

    if 'date_start' in df.columns:
        df['date_start'] = df['date_start'].apply(_format_iso_date)
    if 'date_end' in df.columns:
        df['date_end'] = df['date_end'].apply(_format_iso_date)

    df = df[EXPORT_COLS]
    return df


async def export_to_csv(filters: dict) -> StreamingResponse:
    """Export events to CSV format with UTF-8 BOM and text-formatted phone numbers for Excel compatibility."""
    df = await get_events_df(filters)

    # Format phone number and whatsapp columns for Excel to avoid scientific notation (e.g. 3.55692E+11)
    df_csv = df.copy()
    phone_cols = ['organizer_phone', 'organizer_whatsapp']
    for col in phone_cols:
        if col in df_csv.columns:
            df_csv[col] = df_csv[col].apply(
                lambda v: f'="{v}"' if v and str(v).strip() and str(v).strip().replace('+', '').isdigit() else v
            )

    stream = io.StringIO()
    # Prepend UTF-8 BOM so Excel opens special characters with UTF-8 encoding
    stream.write('\ufeff')
    df_csv.to_csv(stream, index=False)
    stream.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"lmscraper_export_{timestamp}.csv"

    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def export_to_xlsx(filters: dict) -> StreamingResponse:
    """Export events to XLSX format."""
    df = await get_events_df(filters)

    stream = io.BytesIO()

    with pd.ExcelWriter(stream, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Events')
        worksheet = writer.sheets['Events']

        # Header formatting
        header_fill = PatternFill(start_color='1a1c23', end_color='1a1c23', fill_type='solid')
        header_font = Font(color='f5a623', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Style record_id and hidden columns in grey to signal they are system columns
        system_col_font = Font(color='888888', italic=True)
        system_cols = {'record_id', 'hidden'}
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in system_cols:
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.font = system_col_font

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # Auto-fit column widths
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max((df[col].astype(str).map(len).max(), len(col)))
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    stream.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"lmscraper_export_{timestamp}.xlsx"

    return StreamingResponse(
        io.BytesIO(stream.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------

# Fields the user is allowed to set via an import file (maps file column → DB column)
IMPORT_COL_MAP = {
    'title': 'title',
    'date_start': 'date_start',
    'date_end': 'date_end',
    'city': 'city',
    'country': 'country',
    'venue': 'venue',
    'price': 'price',
    'event_type': 'category',
    'dance_style': 'dance_style',
    'organizer_name': 'organizer_name',
    'organizer_email': 'organizer_email',
    'organizer_phone': 'organizer_phone',
    'organizer_instagram': 'organizer_instagram',
    'organizer_facebook': 'organizer_facebook',
    'organizer_tiktok': 'organizer_tiktok',
    'organizer_whatsapp': 'organizer_whatsapp',
    'organizer_youtube': 'organizer_youtube',
    'organizer_twitter': 'organizer_twitter',
    'organizer_website': 'organizer_website',
    'contact_hidden': 'contact_hidden',
    'event_url': 'event_url',
    'platform': 'platform',
    'hidden': 'is_hidden',
}

# Read-only columns that are silently ignored if edited in the file
READ_ONLY_COLS = {'record_id', 'source_domain', 'scraped_at'}

# Fields shown in diff previews
PREVIEW_DISPLAY_FIELDS = ['title', 'city', 'date_start', 'organizer_name', 'event_url']


def parse_import_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Parse a CSV or XLSX upload into a DataFrame.

    Raises ValueError with a descriptive message on bad input.
    """
    fname = filename.lower()
    try:
        if fname.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
        elif fname.endswith('.xlsx') or fname.endswith('.xls'):
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, dtype=str)
            df = df.fillna('')
        else:
            raise ValueError("Unsupported file format. Please upload a .csv or .xlsx file.")
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Could not read file: {exc}") from exc

    if df.empty:
        raise ValueError("The uploaded file contains no data rows.")

    # Normalise column names: strip whitespace, lowercase
    df.columns = [c.strip().lower() for c in df.columns]

    if 'record_id' not in df.columns:
        raise ValueError(
            "The file is missing the 'record_id' column. "
            "Please use a file exported from lmScraper so that records can be matched."
        )

    # Clean Excel formula text wrapping like ="355692..." if present
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: v[2:-1] if isinstance(v, str) and v.startswith('="') and v.endswith('"') else v
        )

    return df


def _normalise_hidden(val: str) -> int:
    """Convert a 'hidden' cell value to 0 or 1."""
    if isinstance(val, str):
        return 1 if val.strip().lower() in ('yes', '1', 'true') else 0
    return 1 if val else 0


def _values_differ(old, new) -> bool:
    """Compare two cell values, treating blanks as equal."""
    old_s = '' if (old is None or (isinstance(old, float) and pd.isna(old))) else str(old).strip()
    new_s = '' if (new is None or (isinstance(new, float) and pd.isna(new))) else str(new).strip()
    return old_s != new_s


async def compute_import_diff(df: pd.DataFrame, mode: str) -> dict:
    """Compute what would change if *df* were imported with the given *mode*.

    mode='partial': Only rows in the file are considered; absent DB rows are untouched.
    mode='full':    Records in DB that have no matching row in the file are marked hidden.

    Returns a structured diff dict consumed by both the preview endpoint and apply().
    """
    # Separate rows with a known record_id from new insertions
    df = df.copy()
    df['record_id'] = df['record_id'].apply(lambda x: x.strip() if isinstance(x, str) else x)

    has_id_mask = df['record_id'].apply(lambda x: str(x).strip().isdigit())
    df_known = df[has_id_mask].copy()
    df_new = df[~has_id_mask].copy()

    known_ids = [int(x) for x in df_known['record_id']]

    # Fetch current DB state for all referenced IDs
    db_events = await get_events_by_ids(known_ids)

    updated = []
    unchanged = []
    not_found = []

    for _, file_row in df_known.iterrows():
        rid = int(file_row['record_id'])
        row_number = _ + 2  # 1-indexed + header row

        if rid not in db_events:
            not_found.append({'record_id': rid, 'row_number': row_number})
            continue

        db_row = db_events[rid]
        field_changes = {}

        for file_col, db_col in IMPORT_COL_MAP.items():
            if file_col not in df.columns:
                continue
            file_val = file_row.get(file_col, '')

            if file_col == 'hidden':
                new_val = _normalise_hidden(file_val)
                old_val = db_row.get('is_hidden', 0) or 0
                if int(new_val) != int(old_val):
                    field_changes['hidden'] = [
                        'yes' if old_val else 'no',
                        'yes' if new_val else 'no',
                    ]
            else:
                old_val = db_row.get(db_col, '') or ''
                
                if file_col in ('date_start', 'date_end'):
                    old_val = _normalize_event_date(old_val)
                    file_val = _normalize_event_date(file_val)
                    
                if file_col == 'contact_hidden':
                    file_val = str(_normalise_hidden(file_val))
                    old_val = str(_normalise_hidden(old_val))
                    
                if _values_differ(old_val, file_val):
                    field_changes[file_col] = [str(old_val), str(file_val)]

        if field_changes:
            updated.append({
                'record_id': rid,
                'title': db_row.get('title', ''),
                'field_changes': field_changes,
            })
        else:
            unchanged.append({'record_id': rid})

    # New insertions (rows with blank record_id)
    inserted = []
    for _, file_row in df_new.iterrows():
        preview = {fc: str(file_row.get(fc, '')) for fc in PREVIEW_DISPLAY_FIELDS if fc in df.columns}
        inserted.append(preview)

    # Full mode: records in DB not in the file → mark hidden
    removed = []
    if mode == 'full':
        all_db_ids = await get_all_event_ids()
        file_id_set = set(known_ids)
        for missing_id in all_db_ids - file_id_set:
            db_row = db_events.get(missing_id)
            if db_row is None:
                # Fetch individually if not already loaded
                fetched = await get_events_by_ids([missing_id])
                db_row = fetched.get(missing_id, {})
            # Only mark as removed if currently visible
            if not db_row.get('is_hidden', 0):
                removed.append({
                    'record_id': missing_id,
                    'title': db_row.get('title', ''),
                })

    return {
        'mode': mode,
        'summary': {
            'updated': len(updated),
            'inserted': len(inserted),
            'removed': len(removed),
            'unchanged': len(unchanged),
            'not_found': len(not_found),
        },
        'rows': {
            'updated': updated,
            'inserted': inserted,
            'removed': removed,
            'not_found': not_found,
        },
        # Internal data needed by apply_import (not sent to frontend in preview)
        '_df': df,
        '_db_events': db_events,
        '_known_ids': known_ids,
    }


async def apply_import(diff: dict) -> dict:
    """Apply a pre-computed import diff to the database.

    Returns counts of applied changes.
    """
    df: pd.DataFrame = diff['_df']
    db_events: dict = diff['_db_events']
    mode: str = diff['mode']

    updates = []

    # Build update payloads for ONLY the changed rows
    has_id_mask = df['record_id'].apply(lambda x: str(x).strip().isdigit())
    df_known = df[has_id_mask].copy()
    
    updated_ids = {r['record_id'] for r in diff['rows']['updated']}

    for _, file_row in df_known.iterrows():
        rid = int(file_row['record_id'])
        if rid not in updated_ids:
            continue  # unchanged or not_found — skip

        update = {'id': rid}
        for file_col, db_col in IMPORT_COL_MAP.items():
            if file_col not in df.columns:
                continue
            file_val = file_row.get(file_col, '')
            if file_col == 'hidden':
                update['is_hidden'] = _normalise_hidden(file_val)
            else:
                update[db_col] = str(file_val).strip() if file_val != '' else None
        updates.append(update)

    updated_count = await bulk_update_events(updates) if updates else 0

    # Full mode: hide records absent from the file
    removed_count = 0
    if mode == 'full' and diff['rows']['removed']:
        hide_updates = [
            {'id': r['record_id'], 'is_hidden': 1}
            for r in diff['rows']['removed']
        ]
        removed_count = await bulk_update_events(hide_updates)

    # Insert new rows (blank record_id)
    df_new = df[~has_id_mask].copy()
    inserted_count = 0
    import uuid as _uuid
    from datetime import datetime as _dt

    for _, file_row in df_new.iterrows():
        event_dict = {}
        for file_col, db_col in IMPORT_COL_MAP.items():
            if file_col == 'hidden' or file_col not in df.columns:
                continue
            val = str(file_row.get(file_col, '')).strip()
            event_dict[db_col] = val if val else None

        event_dict['job_id'] = 'manual_import'
        event_dict['scraped_at'] = _dt.utcnow().isoformat()
        # Generate a unique content hash for manually inserted rows
        import hashlib
        hash_src = f"{event_dict.get('title', '')}|{event_dict.get('date_start', '')}|{event_dict.get('event_url', '')}|{_dt.utcnow().isoformat()}"
        event_dict['content_hash'] = hashlib.md5(hash_src.encode()).hexdigest()
        event_dict['is_hidden'] = _normalise_hidden(file_row.get('hidden', 'no'))

        _id, _status = await insert_event(event_dict)
        if _status in ('new', 'duplicate_updated'):
            inserted_count += 1

    return {
        'updated': updated_count,
        'inserted': inserted_count,
        'removed': removed_count,
    }
